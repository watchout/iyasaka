from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "dev-os_exit_plan_2026.xlsx"
NAVY = "1A365D"
ACCENT = "FF9E00"
HF = PatternFill("solid", fgColor="D9E1F2")
FT = Font(name="Calibri", size=16, bold=True, color=NAVY)
FH = Font(name="Calibri", size=11, bold=True, color="0F172A")
FM = Font(name="Calibri", size=11, color="6B7280")
TH = Side(style="thin", color="B4C6E7")
BA = Border(top=TH, bottom=TH, left=TH, right=TH)
AW = Alignment(vertical="top", wrap_text=True)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)

def sh(ws, row, nc):
    for c in range(1, nc+1):
        cl = ws.cell(row=row, column=c)
        cl.font, cl.fill, cl.alignment, cl.border = FH, HF, AC, BA

def sw(ws, ww):
    for i, w in enumerate(ww, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def bb(ws, sr, er, nc):
    for r in range(sr, er+1):
        for c in range(1, nc+1):
            ws.cell(row=r, column=c).border = BA
            ws.cell(row=r, column=c).alignment = AW

wb = Workbook()
ws = wb.active
ws.title = "README"
ws.sheet_properties.tabColor = NAVY
for r in [["dev-OS 2026 売却実行計画"],[],["項目","内容"],["売り方","事業譲渡（顧客付き）"],["売却後の権利","IYASAKA 永久ライセンスバック"],["KGI","2026年内クローズ / 譲渡対価 1億 / 永久ライセンスバック"],["北極星","NDA締結済み買い手数 / 承継可能MRR"],["WBR","30分: スコア更新 → 未達原因1つ → 打ち手3つ以内"],["MBR","60分: MRR/顧客/買い手/DDリスク棚卸し"]]:
    ws.append(r)
ws.cell(1,1).font = FT
ws.merge_cells("A1:B1")
sh(ws, 3, 2)
bb(ws, 4, 9, 2)
sw(ws, [22, 80])

ws2 = wb.create_sheet("KPI定義")
ws2.sheet_properties.tabColor = "3182CE"
kh = ["KPI_ID","KPI名","区分","定義","頻度","目標","オーナー"]
ws2.append(kh)
sh(ws2, 1, len(kh))
for rd in [["KGI-1","クローズ（入金）","KGI","2026年内に入金完了","月次","Yes","CEO"],["KGI-2","譲渡対価","KGI","前金+アーンアウト合計","月次",100000000,"CEO"],["KGI-3","永久ライセンスバック","KGI","非独占/取消不能の永久利用条項","月次","Yes","法務/CEO"],["NS-1","NDA締結済み買い手数","北極星","実名交渉できる買い手数","週次",8,"BizDev"],["NS-2","承継可能MRR","北極星","譲渡時に引き継げるMRR合計","週次",2000000,"CS/経理"],["W-1","新規打診数","週次","今週新たに接触した買い手数","週次",20,"BizDev"],["W-2","NDA獲得数","週次","今週NDA締結した件数","週次","1-2","BizDev"],["W-3","買い手面談数","週次","意思決定者参加の面談数","週次",2,"CEO/BizDev"],["W-4","LOI進行数","週次","LOI依頼済み/交渉中の社数","週次","1+","CEO"],["W-5","DR完成度","週次","DD資料が揃っている割合","週次","100%","法務/CTO"]]:
    ws2.append(rd)
bb(ws2, 2, 11, len(kh))
for row in ws2.iter_rows(min_row=2, max_col=len(kh)):
    if isinstance(row[5].value, (int,float)) and row[5].value > 1000:
        row[5].number_format = '#,##0'
sw(ws2, [9,22,10,40,8,16,12])

ws3 = wb.create_sheet("週次WBR")
ws3.sheet_properties.tabColor = ACCENT
wh = ["WeekStart","Wk","打診T","打診A","NDA_T","NDA_A","面談T","面談A","LOI_T","LOI_A","有償T","有償A","承継MRR_T","承継MRR_A","DR_T","DR_A","Blk_T","Blk_A","ボトルネック","打ち手","担当"]
ws3.append(wh)
sh(ws3, 1, len(wh))
s = date(2026, 2, 9)
wt = [(10,1,1,0,0,0,0.3,3),(10,1,1,0,0,0,0.6,2),(20,1,2,0,1,500000,1.0,2),(20,1,2,1,1,500000,1.0,1),(20,1,2,1,2,1500000,1.0,1),(20,1,2,1,2,1500000,1.0,1),(20,1,2,1,2,1500000,1.0,1),(20,1,2,1,2,1500000,1.0,1),(20,1,2,1,2,2000000,1.0,1),(20,1,2,1,2,2000000,1.0,1),(20,1,2,1,2,2000000,1.0,1),(20,1,2,1,3,2000000,1.0,1)]
for i,(o,n,m,l,p,mrr,dr,b) in enumerate(wt):
    dt = s + timedelta(days=7*i)
    ws3.append([dt,i+1,o,None,n,None,m,None,l,None,p,None,mrr,None,dr,None,b,None,None,None,None])
for row in ws3.iter_rows(min_row=2, max_col=len(wh)):
    row[0].number_format = 'yyyy-mm-dd'
    row[12].number_format = '#,##0'
    row[13].number_format = '#,##0'
    row[14].number_format = '0%'
    row[15].number_format = '0%'
bb(ws3, 2, 1+len(wt), len(wh))
sw(ws3, [12,5,8,8,8,8,8,8,8,8,8,8,14,14,8,8,10,8,32,36,14])
ws3.freeze_panes = "A2"

ws4 = wb.create_sheet("月次MBR")
ws4.sheet_properties.tabColor = "319795"
mh = ["Month","承継MRR_T","承継MRR_A","ARR_T","ARR_A","有償T","有償A","Top1集中T","Top1集中A","NDA累計T","NDA累計A","LOI累計T","LOI累計A","DD社数","Close%_T","Close%_A","学び","リスク","翌月最優先"]
ws4.append(mh)
sh(ws4, 1, len(mh))
for rd in [["2026-02",500000,None,6000000,None,1,None,0.8,None,2,None,0,None,0,0,None,None,None,None],["2026-03",1500000,None,18000000,None,2,None,0.6,None,5,None,1,None,0,0.1,None,None,None,None],["2026-04",2000000,None,24000000,None,2,None,0.6,None,8,None,1,None,1,0.2,None,None,None,None],["2026-05",2000000,None,24000000,None,3,None,0.5,None,10,None,2,None,1,0.3,None,None,None,None],["2026-06",2000000,None,24000000,None,3,None,0.5,None,10,None,2,None,1,0.4,None,None,None,None],["2026-07",2500000,None,30000000,None,4,None,0.4,None,12,None,2,None,2,0.5,None,None,None,None],["2026-08",2500000,None,30000000,None,4,None,0.4,None,12,None,2,None,2,0.6,None,None,None,None],["2026-09",3000000,None,36000000,None,5,None,0.3,None,14,None,3,None,2,0.7,None,None,None,None],["2026-10",3000000,None,36000000,None,5,None,0.3,None,14,None,3,None,2,0.8,None,None,None,None],["2026-11",3000000,None,36000000,None,5,None,0.3,None,15,None,3,None,2,0.9,None,None,None,None],["2026-12",3000000,None,36000000,None,5,None,0.3,None,15,None,3,None,2,1.0,None,None,None,None]]:
    ws4.append(rd)
for row in ws4.iter_rows(min_row=2, max_col=len(mh)):
    for ci in (1,2,3,4): row[ci].number_format = '#,##0'
    for ci in (7,8,14,15): row[ci].number_format = '0%'
bb(ws4, 2, 12, len(mh))
sw(ws4, [10,14,14,16,16,8,8,10,10,10,10,10,10,8,10,10,28,28,32])
ws4.freeze_panes = "A2"

ws5 = wb.create_sheet("買い手パイプライン")
ws5.sheet_properties.tabColor = "38A169"
ph = ["買い手名","カテゴリ","Fit","チャネル","担当者","打診日","NDA日","面談日","ステージ","確度","論点","次アクション","オーナー","備考"]
ws5.append(ph)
sh(ws5, 1, len(ph))
for cat in ["SI/受託","開発ツール系","SI/受託","事業会社(情シス)","SI/受託"]:
    ws5.append([None,cat,"A","直打診",None,None,None,None,"Prospect",0.1,"永久ライセンスバック必須",None,None,None])
for row in ws5.iter_rows(min_row=2, max_col=len(ph)):
    row[5].number_format = 'yyyy-mm-dd'
    row[6].number_format = 'yyyy-mm-dd'
    row[7].number_format = 'yyyy-mm-dd'
    row[9].number_format = '0%'
bb(ws5, 2, 6, len(ph))
sw(ws5, [22,14,5,14,18,12,12,12,12,8,28,22,14,22])
ws5.freeze_panes = "A2"

ws6 = wb.create_sheet("顧客一覧")
ws6.sheet_properties.tabColor = "3182CE"
ch = ["顧客名","状態","契約形態","開始日","MRR","承継","更新/終了","ユースケース","備考"]
ws6.append(ch)
sh(ws6, 1, len(ch))
for _ in range(5):
    ws6.append([None,"Pilot","月額",None,None,"要同意",None,None,None])
for row in ws6.iter_rows(min_row=2, max_col=len(ch)):
    row[3].number_format = 'yyyy-mm-dd'
    row[4].number_format = '#,##0'
bb(ws6, 2, 6, len(ch))
sw(ws6, [22,10,12,12,14,10,12,32,28])

ws7 = wb.create_sheet("データルーム")
ws7.sheet_properties.tabColor = "E53E3E"
dh = ["領域","項目","担当","状態","リンク/パス","備考"]
ws7.append(dh)
sh(ws7, 1, len(dh))
for a,it,ow in [("IP","ソースコード権利関係","法務"),("契約","顧客契約（承継同意条項）","法務"),("契約","永久ライセンスバック条項","法務"),("財務","dev-OS単体PL","経理"),("顧客","顧客一覧","CS"),("プロダクト","機能一覧/ロードマップ","CTO"),("運用","運用手順書","CTO/CS"),("セキュリティ","権限/鍵/監査ログ","CTO"),("法務","NDA/LOI/SPAテンプレ","法務"),("営業","Teaser/CIM/FAQ","BizDev")]:
    ws7.append([a,it,ow,"Not started",None,None])
bb(ws7, 2, 11, len(dh))
sw(ws7, [12,44,12,14,28,28])

ws8 = wb.create_sheet("12週ロードマップ")
ws8.sheet_properties.tabColor = NAVY
rh = ["Week","フォーカス","成果物","KPIターゲット","担当","注記"]
ws8.append(rh)
sh(ws8, 1, len(rh))
for rd in [(1,"DD耐性の土台","事業境界図/IP棚卸し","DR30%/打診10","法務/CTO","重大ブロッカーだけ"),(2,"DD耐性(続)","契約雛形/運用手順/Q&A","DR60%/NDA1","法務/BizDev","NDA雛形改善"),(3,"顧客付き商品化","導入パッケージ/デモ","有償1/面談2","CTO/CS","有償化を先に"),(4,"買い手打診","Teaser/CIM/面談","NDA5/LOI候補1","CEO/BizDev","2社並走"),(5,"顧客契約化","承継可能契約へ","有償2/承継MRR150万","CS/法務","NG条項即修正"),(6,"条件交渉","条件表テンプレ","LOI候補1","CEO","永久ライセンス必須"),(7,"DD準備","DDリスト/Q&A","DR100%","法務/CTO","台帳に集約"),(8,"DD進行","面談/Q&A更新","NDA8/DD開始","CEO/BizDev","競争状態維持"),(9,"最終条件","SPA/TSA","Close20%","法務/CEO","リスク潰す"),(10,"Close準備","承継同意回収","Close40%","CS/法務","同意遅延=最大リスク"),(11,"移行","引継ぎ/運用分離","Close60%","CTO/CS","責任分界明確"),(12,"クロージング","締結/入金/引継ぎ","クローズ","CEO","永久ライセンス最終確認")]:
    ws8.append(list(rd))
bb(ws8, 2, 13, len(rh))
sw(ws8, [6,20,36,28,14,28])

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(OUT))
print(f"OK: {OUT}")
