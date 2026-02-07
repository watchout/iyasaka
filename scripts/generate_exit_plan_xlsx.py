"""dev-OS 2026 Exit Plan Excel Generator"""
from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "dev-os_exit_plan_2026.xlsx"
NAVY = "1A365D"
ACCENT = "FF9E00"
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=NAVY)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="0F172A")
FONT_MUTED = Font(name="Calibri", size=11, color="6B7280")
THIN = Side(style="thin", color="B4C6E7")
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
ALIGN_WRAP = Alignment(vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
YEN_FMT = '#,##0'
PCT_FMT = '0%'
DATE_FMT = 'yyyy-mm-dd'

def style_hdr(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL

def set_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def borders(ws, sr, er, nc):
    for r in range(sr, er + 1):
        for c in range(1, nc + 1):
            ws.cell(row=r, column=c).border = BORDER_ALL
            ws.cell(row=r, column=c).alignment = ALIGN_WRAP

wb = Workbook()

# README
ws = wb.active
ws.title = "README"
ws.sheet_properties.tabColor = NAVY
readme = [
    ["dev-OS 2026 売却実行計画"],
    ["Weekly Business Review / Monthly Business Review"],
    [],
    ["項目", "内容"],
    ["売り方", "事業譲渡（顧客付き）"],
    ["売却後の権利", "IYASAKA 永久ライセンスバック（必須/非独占/取消不能/内部利用）"],
    ["KGI", "2026年内クローズ（入金）/ 譲渡対価 1億 / 永久ライセンスバック条項"],
    ["北極星（週次）", "NDA締結済み買い手数 / 承継可能な有償MRR"],
    ["週次WBR", "30分: スコア更新 → 未達の原因を1つ → 打ち手は3つ以内"],
    ["月次MBR", "60分: MRR/顧客/買い手/DDリスク棚卸し → 翌月ターゲット更新"],
]
for r in readme:
    ws.append(r)
ws.cell(1,1).font = FONT_TITLE
ws.cell(2,1).font = FONT_MUTED
ws.merge_cells("A1:B1")
ws.merge_cells("A2:B2")
style_hdr(ws, 4, 2)
borders(ws, 5, 10, 2)
set_w(ws, [22, 80])

# KPI
ws2 = wb.create_sheet("KPI定義")
ws2.sheet_properties.tabColor = "3182CE"
kh = ["KPI_ID","KPI名","区分","定義","算出方法","頻度","目標","オーナー","ソース","注記"]
kd = [
    ["KGI-1","クローズ（入金）","KGI","2026年内に売却完了し入金まで確認","Yes/No","月次","Yes","CEO","入金/契約","入金まで"],
    ["KGI-2","譲渡対価","KGI","前金+アーンアウト合計","円","月次",100000000,"CEO","LOI/SPA","条件付き対価は達成条件も記録"],
    ["KGI-3","永久ライセンスバック","KGI","非独占/取消不能で永久利用できる条項","Yes/No","月次","Yes","法務/CEO","TS/LOI/SPA","段階で管理"],
    ["NS-1","NDA締結済み買い手数","北極星","実名交渉できる買い手数","社数","週次",8,"BizDev","パイプライン",""],
    ["NS-2","承継可能MRR","北極星","譲渡時に引き継げる契約のMRR合計","円","週次",2000000,"CS/経理","契約/請求","承継同意が取れる契約のみ"],
    ["W-1","新規打診数","週次","今週新たに接触した買い手数","社数","週次",20,"BizDev","アウトリーチ",""],
    ["W-2","NDA獲得数","週次","今週NDA締結した件数","社数","週次","1-2","BizDev","法務/メール",""],
    ["W-3","買い手面談数","週次","意思決定者参加の面談回数","回","週次",2,"CEO/BizDev","議事録",""],
    ["W-4","LOI進行数","週次","LOI依頼済み/交渉中の社数","社数","週次","1+","CEO","パイプライン","2社並走"],
    ["W-5","データルーム完成度","週次","DD資料が揃っている割合","%","週次","100%","法務/CTO","データルーム",""],
]
ws2.append(kh)
style_hdr(ws2, 1, len(kh))
for rd in kd:
    ws2.append(rd)
borders(ws2, 2, 1+len(kd), len(kh))
for row in ws2.iter_rows(min_row=2, max_col=len(kh)):
    if isinstance(row[6].value, (int,float)) and row[6].value > 1000:
        row[6].number_format = YEN_FMT
set_w(ws2, [9,22,10,40,10,8,16,12,14,32])

# WBR
ws3 = wb.create_sheet("週次WBR")
ws3.sheet_properties.tabColor = ACCENT
wh = ["WeekStart","Wk","打診T","打診A","NDA T","NDA A","面談T","面談A","LOI T","LOI A","有償T","有償A","承継MRR T","承継MRR A","DR T","DR A","BlkMax T","Blk A","ボトルネック","打ち手","担当"]
ws3.append(wh)
style_hdr(ws3, 1, len(wh))
start = date(2026, 2, 9)
wt = [
    (10,1,1,0,0,0,0.3,3),(10,1,1,0,0,0,0.6,2),(20,1,2,0,1,500000,1.0,2),
    (20,1,2,1,1,500000,1.0,1),(20,1,2,1,2,1500000,1.0,1),(20,1,2,1,2,1500000,1.0,1),
    (20,1,2,1,2,1500000,1.0,1),(20,1,2,1,2,1500000,1.0,1),(20,1,2,1,2,2000000,1.0,1),
    (20,1,2,1,2,2000000,1.0,1),(20,1,2,1,2,2000000,1.0,1),(20,1,2,1,3,2000000,1.0,1),
]
for i,(o,n,m,l,p,mrr,dr,b) in enumerate(wt):
    dt = start + timedelta(days=7*i)
    ws3.append([dt,i+1,o,None,n,None,m,None,l,None,p,None,mrr,None,dr,None,b,None,None,None,None])
for row in ws3.iter_rows(min_row=2, max_col=len(wh)):
    row[0].number_format = DATE_FMT
    row[12].number_format = YEN_FMT
    row[13].number_format = YEN_FMT
    row[14].number_format = PCT_FMT
    row[15].number_format = PCT_FMT
borders(ws3, 2, 1+len(wt), len(wh))
set_w(ws3, [12,5,8,8,8,8,8,8,8,8,8,8,14,14,8,8,10,8,32,36,14])
ws3.freeze_panes = "A2"

# MBR
ws4 = wb.create_sheet("月次MBR")
ws4.sheet_properties.tabColor = "319795"
mh = ["Month","承継MRR T","承継MRR A","ARR T","ARR A","有償T","有償A","Top1集中T","Top1集中A","NDA累計T","NDA累計A","LOI累計T","LOI累計A","DD社数","Close% T","Close% A","学び/所感","主要リスク","翌月の最優先"]
ws4.append(mh)
style_hdr(ws4, 1, len(mh))
md = [
    ["2026-02",500000,None,6000000,None,1,None,0.8,None,2,None,0,None,0,0,None,None,None,None],
    ["2026-03",1500000,None,18000000,None,2,None,0.6,None,5,None,1,None,0,0.1,None,None,None,None],
    ["2026-04",2000000,None,24000000,None,2,None,0.6,None,8,None,1,None,1,0.2,None,None,None,None],
    ["2026-05",2000000,None,24000000,None,3,None,0.5,None,10,None,2,None,1,0.3,None,None,None,None],
    ["2026-06",2000000,None,24000000,None,3,None,0.5,None,10,None,2,None,1,0.4,None,None,None,None],
    ["2026-07",2500000,None,30000000,None,4,None,0.4,None,12,None,2,None,2,0.5,None,None,None,None],
    ["2026-08",2500000,None,30000000,None,4,None,0.4,None,12,None,2,None,2,0.6,None,None,None,None],
    ["2026-09",3000000,None,36000000,None,5,None,0.3,None,14,None,3,None,2,0.7,None,None,None,None],
    ["2026-10",3000000,None,36000000,None,5,None,0.3,None,14,None,3,None,2,0.8,None,None,None,None],
    ["2026-11",3000000,None,36000000,None,5,None,0.3,None,15,None,3,None,2,0.9,None,None,None,None],
    ["2026-12",3000000,None,36000000,None,5,None,0.3,None,15,None,3,None,2,1.0,None,None,None,None],
]
for rd in md:
    ws4.append(rd)
for row in ws4.iter_rows(min_row=2, max_col=len(mh)):
    for ci in (1,2,3,4): row[ci].number_format = YEN_FMT
    for ci in (7,8,14,15): row[ci].number_format = PCT_FMT
borders(ws4, 2, 1+len(md), len(mh))
set_w(ws4, [10,14,14,16,16,8,8,10,10,10,10,10,10,8,10,10,28,28,32])
ws4.freeze_panes = "A2"

# Pipeline
ws5 = wb.create_sheet("買い手パイプライン")
ws5.sheet_properties.tabColor = "38A169"
ph = ["買い手名","カテゴリ","Fit","チャネル","担当者/役職","打診日","NDA日","面談日","ステージ","確度","論点/条件","次アクション","オーナー","備考"]
ws5.append(ph)
style_hdr(ws5, 1, len(ph))
cats = ["SI/受託","開発ツール系","SI/受託","事業会社(情シス)","SI/受託"]
for cat in cats:
    ws5.append([None,cat,"A","直打診",None,None,None,None,"Prospect",0.1,"永久ライセンスバック必須",None,None,None])
for row in ws5.iter_rows(min_row=2, max_col=len(ph)):
    row[5].number_format = DATE_FMT
    row[6].number_format = DATE_FMT
    row[7].number_format = DATE_FMT
    row[9].number_format = PCT_FMT
borders(ws5, 2, 1+len(cats), len(ph))
set_w(ws5, [22,14,5,14,18,12,12,12,12,8,28,22,14,22])
ws5.freeze_panes = "A2"

# Customer
ws6 = wb.create_sheet("顧客一覧")
ws6.sheet_properties.tabColor = "3182CE"
ch = ["顧客名","状態","契約形態","開始日","MRR","承継","更新/終了","ユースケース","備考"]
ws6.append(ch)
style_hdr(ws6, 1, len(ch))
for _ in range(5):
    ws6.append([None,"Pilot","月額",None,None,"要同意",None,None,None])
for row in ws6.iter_rows(min_row=2, max_col=len(ch)):
    row[3].number_format = DATE_FMT
    row[4].number_format = YEN_FMT
    row[6].number_format = DATE_FMT
borders(ws6, 2, 6, len(ch))
set_w(ws6, [22,10,12,12,14,10,12,32,28])

# Data Room
ws7 = wb.create_sheet("データルーム")
ws7.sheet_properties.tabColor = "E53E3E"
dh = ["領域","項目","担当","状態","リンク/パス","備考"]
ws7.append(dh)
style_hdr(ws7, 1, len(dh))
di = [
    ("IP","ソースコード権利関係","法務"),
    ("契約","顧客契約（承継同意条項）","法務"),
    ("契約","永久ライセンスバック条項（雛形）","法務"),
    ("財務","dev-OS単体の売上/費用（簡易PL）","経理"),
    ("顧客","顧客一覧（契約/単価/更新/承継可否）","CS"),
    ("プロダクト","機能一覧/ロードマップ","CTO"),
    ("運用","運用手順書（導入/保守/障害対応）","CTO/CS"),
    ("セキュリティ","権限/鍵/監査ログ","CTO"),
    ("法務","NDA/LOI/SPAテンプレ","法務"),
    ("営業","Teaser/CIM/FAQ","BizDev"),
]
for area,item,owner in di:
    ws7.append([area,item,owner,"Not started",None,None])
borders(ws7, 2, 1+len(di), len(dh))
set_w(ws7, [12,44,12,14,28,28])

# Roadmap
ws8 = wb.create_sheet("12週ロードマップ")
ws8.sheet_properties.tabColor = NAVY
rh = ["Week","フォーカス","成果物","KPIターゲット","担当","注記"]
ws8.append(rh)
style_hdr(ws8, 1, len(rh))
rm = [
    (1,"DD耐性の土台","事業境界図、IP棚卸し","DR 30%, 打診10","法務/CTO","重大ブロッカーだけ数える"),
    (2,"DD耐性の土台(続)","契約雛形、運用手順、Q&A台帳","DR 60%, NDA 1","法務/BizDev","NDA雛形改善"),
    (3,"顧客付き商品化","導入パッケージ、デモ環境","有償パイロット1, 面談2","CTO/CS","有償化を先に"),
    (4,"買い手打診・面談","Teaser/CIM、意思決定者面談","NDA累計5, LOI候補1","CEO/BizDev","2社並走を作る"),
    (5,"顧客を契約にする","承継可能契約へ寄せる","有償2, 承継MRR 150万","CS/法務","承継NG条項は即修正"),
    (6,"条件交渉","条件表テンプレ化","LOI候補1","CEO","永久ライセンスは必須"),
    (7,"DD準備","DDチェックリスト、Q&A整備","DR 100%維持","法務/CTO","追加質問は台帳に集約"),
    (8,"DD進行","追加面談、Q&A更新","NDA累計8, DD開始","CEO/BizDev","競争状態を維持"),
    (9,"最終条件","SPAドラフト、TSA確定","クロージング20%","法務/CEO","契約リスクを潰す"),
    (10,"クロージング準備","顧客承継同意回収","クロージング40%","CS/法務","同意遅延が最大リスク"),
    (11,"移行","引継ぎ、運用レーン分離","クロージング60%","CTO/CS","責任分界を明確に"),
    (12,"クロージング","締結/入金/引継ぎ完了","クローズ","CEO","永久ライセンスの最終確認"),
]
for rd in rm:
    ws8.append(list(rd))
borders(ws8, 2, 1+len(rm), len(rh))
set_w(ws8, [6,20,44,28,14,32])

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(OUT))
print(f"OK: {OUT}")
